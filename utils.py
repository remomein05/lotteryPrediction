"""Shared utility functions for number processing and file operations."""

from __future__ import annotations

import math
import os
import sys
import time
from collections import Counter
from typing import Any, Dict, List, Set, Tuple

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("Error: The 'openpyxl' library is required to read Excel files.")
    print("Please install it by running: pip install openpyxl")
    sys.exit(1)


# Global cache for block origins to speed up processing
_BLOCK_ORIGIN_CACHE: Dict[Tuple[int, int], Tuple[int, int]] = {}

# Cache for transformation results to optimize get_block()
_TRANSFORMATION_CACHE: Dict[Tuple[str, str], str] = {}


def _apply_transformation(number_str: str, operation: str) -> str:
    """Apply a transformation operation to a number string with caching."""
    cache_key = (number_str, operation)
    if cache_key in _TRANSFORMATION_CACHE:
        return _TRANSFORMATION_CACHE[cache_key]
    
    if operation == "base":
        result = number_str
    elif operation == "sb":
        result = "".join(sorted([str((10 - int(d)) % 10) for d in number_str], reverse=True))
    elif operation == "pot":
        result = "".join(sorted([str((int(d) * 3) % 10) for d in number_str], reverse=True))
    elif operation == "fa":
        result = "".join(sorted([str((int(d) + 5) % 10) for d in number_str], reverse=True))
    else:
        raise ValueError(f"Unknown operation: {operation}")
    
    _TRANSFORMATION_CACHE[cache_key] = result
    return result


def calculate_permutations(number_str: str) -> int:
    """Calculates the number of unique permutations for a given string of digits."""
    if not number_str or not number_str.isdigit():
        return 0
    counts = Counter(number_str)
    perms = math.factorial(len(number_str))
    for count in counts.values():
        perms //= math.factorial(count)
    return perms


def get_sb(number_str: str) -> str:
    """Applies the SB logic: (10 - digit) % 10 for each digit, sorted descending."""
    return _apply_transformation(number_str, "sb")


def get_pot(number_str: str) -> str:
    """Applies the POT logic: (digit * 3) % 10 for each digit, sorted descending."""
    return _apply_transformation(number_str, "pot")


def get_fa(number_str: str) -> str:
    """Applies the FA logic: (digit + 5) % 10 for each digit, sorted descending."""
    return _apply_transformation(number_str, "fa")


def get_family(number_str: str) -> List[str]:
    """Generates a 'Family' of 4 numbers by adding 1 to each digit (modulo 10) 4 times."""
    family: List[str] = []
    current = number_str
    for _ in range(4):
        next_num = "".join([str((int(d) + 1) % 10) for d in current])
        family.append(next_num)
        current = next_num
    return family


def get_block(base_val: str) -> Dict[str, str]:
    """Generates the 8-member Block for a given base number string.
    
    Optimized to cache intermediate transformations to avoid redundant calculations.
    """
    sb_val = get_sb(base_val)
    pot_val = get_pot(base_val)
    fa_val = get_fa(base_val)
    
    return {
        "B": base_val,
        "SB": sb_val,
        "POT": pot_val,
        "POT SB": get_sb(pot_val),
        "FA": fa_val,
        "FA SB": get_sb(fa_val),
        "FA POT": get_pot(fa_val),
        "FA POT SB": get_sb(get_pot(fa_val))
    }


def get_neighbours(number_str: str) -> List[str]:
    """Generates neighbours based on digit repetition and position."""
    neighbours: List[str] = []

    # Rule: If first 2 digits are identical, act on the last 2 digits (indices 2 and 3)
    if len(number_str) >= 4 and number_str[0] == number_str[1]:
        indices_to_modify = [2, 3]
    else:
        # Standard Rule: Act on non-repeating digits
        counts = Counter(number_str)
        indices_to_modify = [i for i, digit in enumerate(number_str) if counts[digit] == 1]

    for i in indices_to_modify:
        val = int(number_str[i])

        for delta in [1, -1]:
            new_digits = list(number_str)
            new_digits[i] = str((val + delta) % 10)
            combined = "".join(sorted(new_digits, reverse=True))
            if combined not in neighbours:
                neighbours.append(combined)

    return neighbours


def get_block_origin(ws: Any, cell_row: int, cell_col: int) -> Tuple[int, int]:
    """Finds the (start_row, start_col) of the number block containing the given cell with caching."""
    if (cell_row, cell_col) in _BLOCK_ORIGIN_CACHE:
        return _BLOCK_ORIGIN_CACHE[(cell_row, cell_col)]

    # Find start_row (scan up until 2 empty rows or row 7)
    start_row: int = cell_row
    while start_row > 7:
        cell_above_1 = ws.cell(row=start_row - 1, column=cell_col)  # type: ignore
        val_above_1 = cell_above_1.value
        is_empty_1 = val_above_1 is None or str(val_above_1).strip() == ""
        if is_empty_1:
            if (start_row - 2) < 7:  # type: ignore
                break
            cell_above_2 = ws.cell(row=start_row - 2, column=cell_col)  # type: ignore
            val_above_2 = cell_above_2.value
            if val_above_2 is None or str(val_above_2).strip() == "":
                break  # 2 consecutive empty rows
        start_row -= 1

    # Find start_col (scan left until empty column or column D=4)
    start_col: int = cell_col
    while start_col > 4:
        cell_left = ws.cell(row=int(start_row), column=int(start_col) - 1)  # type: ignore
        val_left = cell_left.value
        if val_left is None or str(val_left).strip() == "":
            break
        start_col -= 1  # type: ignore

    _BLOCK_ORIGIN_CACHE[(cell_row, cell_col)] = (int(start_row), int(start_col))
    return int(start_row), int(start_col)


def get_company_offset(ws: Any, cell_row: int, cell_col: int) -> int:
    """Finds the company offset (0 to 6) of a given cell within its block."""
    start_row, _ = get_block_origin(ws, cell_row, cell_col)
    return int(int(cell_row) - int(start_row))


def open_file(filepath: str) -> bool:
    """Opens a file with the default application for the current platform."""
    try:
        if sys.platform == "win32":
            os.startfile(filepath)  # type: ignore
        elif sys.platform == "darwin":  # macOS
            import subprocess

            subprocess.Popen(["open", filepath])
        else:  # Linux and other Unix-like systems
            import subprocess

            try:
                # Try libreoffice first with --norestore to avoid recovery dialogs
                subprocess.Popen(
                    ["libreoffice", "--norestore", filepath],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
            except FileNotFoundError:
                # Fall back to xdg-open if libreoffice is not installed
                subprocess.Popen(
                    ["xdg-open", filepath],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
        return True
    except Exception as e:
        print(f"Note: Could not automatically open file: {e}")
        return False


def close_file(filename_pattern: str) -> bool:
    """Closes LibreOffice instances by targeting specific file patterns."""
    try:
        import subprocess

        # Extract just the filename without path for matching
        if os.path.isabs(filename_pattern):
            file_name = os.path.basename(filename_pattern)
        else:
            file_name = filename_pattern

        if not file_name.lower().endswith(".xlsx"):
            file_name = file_name + ".xlsx"

        # On Linux, use pkill to close LibreOffice with this specific file
        if sys.platform != "win32" and sys.platform != "darwin":
            subprocess.run(
                ["pkill", "-f", f"libreoffice.*{file_name}"], capture_output=True
            )
            time.sleep(0.5)
        elif sys.platform == "win32":
            subprocess.run(
                ["taskkill", "/F", "/FI", f"WINDOWTITLE eq *{file_name}*"],
                capture_output=True,
            )

        return True
    except Exception as e:
        print(f"Note: Could not close file: {e}")
        return False


def get_libreoffice_lockfile(filepath: str) -> str:
    """Get the LibreOffice lock file path for a given file."""
    directory = os.path.dirname(os.path.abspath(filepath))
    basename = os.path.basename(filepath)
    return os.path.join(directory, f".~lock.{basename}#")


def is_file_locked_by_libreoffice(filepath: str) -> bool:
    """Check if a file is locked by LibreOffice."""
    return os.path.exists(get_libreoffice_lockfile(filepath))


def wait_for_file_unlock(filepath: str, timeout_seconds: int = 10) -> bool:
    """Wait for a file to be unlocked by LibreOffice."""
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        if not is_file_locked_by_libreoffice(filepath):
            return True
        time.sleep(0.25)
    return False


def get_terminal_color(text: str, hex_color: str) -> str:
    """Returns the text wrapped in ANSI escape codes for a colored background."""
    hex_color = hex_color.lstrip("#")
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)

    # Calculate brightness to determine if text should be black or white for contrast
    brightness = (r * 299 + g * 587 + b * 114) / 1000
    fg_color = "30" if brightness > 128 else "37"  # 30=Black, 37=White

    return f"\033[{fg_color};48;2;{r};{g};{b}m {text} \033[0m"


def clear_cache() -> None:
    """Clear all caches (useful for testing or resetting state)."""
    global _BLOCK_ORIGIN_CACHE, _TRANSFORMATION_CACHE
    _BLOCK_ORIGIN_CACHE.clear()
    _TRANSFORMATION_CACHE.clear()
