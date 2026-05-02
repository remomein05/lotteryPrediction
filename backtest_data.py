import os
import sys
import openpyxl
import re
from typing import cast, Any, List, Set, Tuple
from collections import defaultdict

# Add the script dir to sys.path so we can import from existing modules
script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.append(script_dir)

try:
    from utils import get_neighbours, get_block, get_company_offset, get_block_origin
    from constants import COMPANIES
except ImportError as e:
    print(f"Error importing from utils or constants: {e}")
    sys.exit(1)

def extract_row_from_coord(coord: str) -> int:
    match = re.search(r'\d+', coord)
    if match:
        return int(match.group())
    return 0

def process_backtest():
    backtest_filename = os.path.join(script_dir, "backtest_vellam.xlsx")
    podium_filename = os.path.join(script_dir, "podium.xlsx")
    family_tree_filename = os.path.join(script_dir, "family_tree.xlsx")

    if not all(os.path.exists(f) for f in [backtest_filename, podium_filename, family_tree_filename]):
        print("Error: Missing one or more required Excel files (backtest_vellam.xlsx, podium.xlsx, family_tree.xlsx).")
        return

    print("Loading workbooks...")
    bt_wb = openpyxl.load_workbook(backtest_filename)
    
    # Ensure Sheet 1 is untouched. Copy to Sheet 2 if not exists.
    if "Sheet1" not in bt_wb.sheetnames and len(bt_wb.sheetnames) > 0:
        source_sheet_name = bt_wb.sheetnames[0]
    else:
        source_sheet_name = "Sheet1"
        
    source_ws = bt_wb[source_sheet_name]
    
    # Check if Sheet 2 exists, create it by copying Sheet 1
    if "Sheet2" in bt_wb.sheetnames:
        print("Removing old Sheet2 to start fresh...")
        del bt_wb["Sheet2"]
        
    print(f"Copying '{source_sheet_name}' to 'Sheet2'...")
    target_ws = bt_wb.copy_worksheet(source_ws)
    target_ws.title = "Sheet2"
    
    print("Loading podium.xlsx (data_only)...")
    pd_wb = openpyxl.load_workbook(podium_filename, data_only=True)
    pd_ws = pd_wb.active
    
    print("Loading family_tree.xlsx (data_only)...")
    ft_wb = openpyxl.load_workbook(family_tree_filename, data_only=True)
    ft_ws = ft_wb.active

    row_count = 0
    for row_idx in range(4, target_ws.max_row + 1):
        input_coord = target_ws[f"C{row_idx}"].value
        input_val = target_ws[f"D{row_idx}"].value
        yellow_coord = target_ws[f"E{row_idx}"].value
        yellow_val = target_ws[f"F{row_idx}"].value
        red_coord = target_ws[f"G{row_idx}"].value
        red_val = target_ws[f"H{row_idx}"].value
        
        if not input_coord or not yellow_val or not red_val:
            continue
            
        input_coord = str(input_coord).strip().upper()
        input_val = str(input_val).strip().zfill(4)
        yellow_val = str(yellow_val).strip().zfill(4)
        red_val = str(red_val).strip().zfill(4)
        
        # Determine Day N
        day_n_row = extract_row_from_coord(input_coord)
        if day_n_row == 0:
            print(f"Warning: Could not extract row from {input_coord} on row {row_idx}")
            continue
            
        try:
            pd_input_cell = pd_ws[input_coord]
            target_offset = get_company_offset(pd_ws, pd_input_cell.row, pd_input_cell.column)
        except Exception:
            # Fallback if invalid coord
            target_offset = get_company_offset(pd_ws, day_n_row, 4) # Default to 4
        
        # Step 13 Logic: Find columns in family tree with yellow_val
        columns_with_yellow = set()
        for ft_row in range(3, 47):
            for ft_col in range(3, 26):
                cell = ft_ws.cell(row=ft_row, column=ft_col)
                if cell.value is not None:
                    cval = str(cell.value).strip().zfill(4)
                    if cval == yellow_val:
                        columns_with_yellow.add(ft_col)
                        
        red_neighbours = set(get_neighbours(red_val))
        blue_coords = []
        blue_values = set()
        
        for ft_row in range(3, 47):
            for ft_col in range(3, 26):
                if ft_col in columns_with_yellow:
                    cell = ft_ws.cell(row=ft_row, column=ft_col)
                    if cell.value is not None:
                        cval = str(cell.value).strip().zfill(4)
                        if cval in red_neighbours:
                            blue_coords.append(cell.coordinate)
                            blue_values.add(cval)
                            
        unique_blue_values = sorted(list(blue_values))
        target_ws[f"J{row_idx}"] = ",".join(blue_coords)
        target_ws[f"K{row_idx}"] = ",".join(unique_blue_values)
        
        # Step 14 Logic: Pool generation based on blocks
        pool_M = set() # Blocks of numbers in Col K
        for b_val in unique_blue_values:
            pool_M.update(get_block(b_val).values())
            
        pool_L = set() # Neighbours of all numbers in pool_M
        for block_num in pool_M:
            pool_L.update(get_neighbours(block_num))
            
        # Helper to search podium for pool matches on a specific day
        def check_day(row_index: int, pool: Set[str]) -> List[str]:
            hits = []
            for col_idx in range(4, 19):
                cell = pd_ws.cell(row=row_index, column=col_idx)
                if cell.value is not None and str(cell.value).strip() != "":
                    val_str = str(cell.value).strip().zfill(4)
                    if val_str in pool:
                        hits.append(val_str)
            return hits

        # Day N
        day_n_hits = check_day(day_n_row, pool_L)
        target_ws[f"L{row_idx}"] = ",".join(sorted(set(day_n_hits))) if day_n_hits else "NO"
        
        # Col M: Day N+1, N+2, N+3
        future_hits = []
        
        day_n1_hits = check_day(day_n_row + 1, pool_M)
        if day_n1_hits:
            future_hits.append(f"N+1 - {','.join(sorted(set(day_n1_hits)))}")
            
        day_n2_hits = check_day(day_n_row + 2, pool_M)
        if day_n2_hits:
            future_hits.append(f"N+2 - {','.join(sorted(set(day_n2_hits)))}")
            
        day_n3_hits = check_day(day_n_row + 3, pool_M)
        if day_n3_hits:
            future_hits.append(f"N+3 - {','.join(sorted(set(day_n3_hits)))}")
            
        target_ws[f"M{row_idx}"] = ", ".join(future_hits) if future_hits else ""
        
        # Discard Col N
        target_ws[f"N{row_idx}"] = ""
        
        row_count += 1

    print(f"Saving updated workbook with {row_count} processed rows...")
    bt_wb.save(backtest_filename)
    print("Done! Check 'Sheet2' in 'backtest_vellam.xlsx' for results.")

if __name__ == "__main__":
    process_backtest()
