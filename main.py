from __future__ import annotations

# type: ignore
import os
import sys
import time
from collections import Counter
from typing import Any, cast, List, Set, Tuple, Dict, Optional

from logger import setup_logging
from constants import COMPANIES
from utils import (
    calculate_permutations,
    get_sb,
    get_pot,
    get_fa,
    get_family,
    get_block,
    get_neighbours,
    get_block_origin,
    get_company_offset,
    open_file,
    close_file,
    is_file_locked_by_libreoffice,
    wait_for_file_unlock,
    get_terminal_color,
)

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: The 'openpyxl' library is required to read Excel files.")
    print("Please install it by running: pip install openpyxl")
    sys.exit(1)

# Start dynamically teeing stdout into result.md and hook input()
setup_logging("result.md")


def save_workbook_safely(wb: Any, filename: str) -> None:
    """Saves the workbook and handles PermissionError by prompting the user."""
    while True:
        try:
            wb.save(filename)
            break
        except PermissionError:
            print(f"\n[!] Error: '{filename}' is currently open in another program (like Excel).")
            input("    Please close the file and press ENTER to try saving again...")


def wait_and_close_libreoffice(filename: str, custom_msg: Optional[str] = None) -> None:
    """Closes the given file and waits for LibreOffice to unlock it."""
    close_file(filename)
    basename = os.path.basename(filename)
    if is_file_locked_by_libreoffice(filename):
        if custom_msg:
            print(custom_msg)
        else:
            print(f"Waiting for LibreOffice to release the lock on '{basename}'...")
        if not wait_for_file_unlock(filename, timeout_seconds=10):
            input(f"Please close '{basename}' and press ENTER to continue...")


def step1_and_2_get_initial_cell(ws: Any, excel_filename: str) -> Tuple[str, str, int, str]:
    """Handles Steps 1 and 2: User Cell Selection and Permutation Limitation Check."""
    print("Opening 'podium.xlsx' for reference...")
    open_file(excel_filename)
    
    while True:
        print("\n--- Step 1: Initialization ---")
        cell_coord = input("Enter the cell coordinate to select (e.g., Q1086): ").strip().upper()
        if not cell_coord:
            print("Invalid input.")
            sys.exit(1)
            
        try:
            cell_value = ws[cell_coord].value
        except ValueError:
            print(f"Error: Invalid cell coordinate '{cell_coord}'. Please try again.")
            continue
            
        if cell_value is None:
            print(f"Error: Cell {cell_coord} is empty. Please try again.")
            continue
            
        number = str(cell_value).strip()
        target_offset = get_company_offset(ws, ws[cell_coord].row, ws[cell_coord].column)
        company_name = COMPANIES.get(target_offset, f"Unknown ({target_offset})")
        
        print("\n--- Summary ---")
        print(f"Selected Cell: {cell_coord}")
        print(f"Target Number: {number}")
        print(f"Identified Company: {company_name}")
        
        print("\n--- Step 2: Permutation Limitation Check ---")
        perms = calculate_permutations(number)
        print(f"Number '{number}' generates {perms} unique permutations.")
        
        if perms > 12:
            print(f"Rule Triggered: Number '{number}' has {perms} permutations (must be 12 or less).")
            print("Please select another number with at least one repeating digit.")
            continue
        
        counts = Counter(number)
        repeating_digits = [digit for digit, count in counts.items() if count > 1]
        
        if len(repeating_digits) != 1:
            print(f"Rule Triggered: Number '{number}' does not have exactly one repeating digit.")
            print("Please select another number (e.g., one with a double or triple digit).")
            continue
            
        print("Check passed: Permutations are 12 or less and number has exactly one repeating digit.")
        close_file("podium")
        print("Closed reference 'podium.xlsx' window.")
        return cell_coord, number, target_offset, company_name


def step3_and_4_highlight_neighbours(
    ws: Any, wb: Any, output_filename: str, cell_coord: str, number: str, target_offset: int, company_name: str
) -> List[str]:
    """Handles Steps 3 and 4: Generating Neighbours and Highlighting them."""
    print("\n--- Step 3: Generating Neighbours ---")
    neighbours = get_neighbours(number)
    print(f"Neighbours for '{number}': {', '.join(neighbours)}")
    
    print("\n--- Step 4: Searching and Highlighting ---")
    light_yellow_fill = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid")
    light_green_fill = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid")
    
    found_matches: int = 0
    for row_idx in range(7, ws.max_row + 1):
        for col_idx in range(4, 19): # Columns D to R
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip() != "":
                cell_val_str = str(cell.value).strip().zfill(4)
                if cell_val_str in neighbours and get_company_offset(ws, row_idx, col_idx) == target_offset:
                    cell.fill = light_yellow_fill
                    found_matches += 1
                    
    print(f"Found and highlighted {found_matches} occurrence(s) of Neighbours within '{company_name}'.")
    
    cast(Any, ws)[cell_coord].fill = light_green_fill
    print(f"Highlighted target cell {cell_coord} in light green.")
    
    print(f"Saving results to '{output_filename}'...")
    save_workbook_safely(wb, output_filename)
    
    print("Done! Automatically opening the generated file...")
    open_file(output_filename)
    time.sleep(2)
    return neighbours


def step5_select_highlighted_cell(
    ws: Any, output_filename: str, neighbours: List[str], target_offset: int, original_cell_coord: str
) -> Tuple[str, str, bool]:
    """Handles Step 5: User selection of a highlighted cell."""
    print("\n--- Step 5: Select Highlighted Cell ---")
    print("Please review the highlighted dataset in Excel.")
    
    while True:
        selected_highlight = input("Enter the cell coordinate of one of the highlighted cells (or 'exit' to quit): ").strip().upper()

        if selected_highlight == 'EXIT' or not selected_highlight:
            print("Exiting...")
            return original_cell_coord, "", True

        try:
            sel_cell = cast(Any, ws)[selected_highlight]
            sel_val = str(sel_cell.value).strip() if sel_cell.value is not None else ""
        except ValueError:
            print(f"Error: Invalid cell coordinate '{selected_highlight}'.")
            continue

        if sel_val in neighbours and get_company_offset(ws, int(sel_cell.row), int(sel_cell.column)) == target_offset: # type: ignore
            wait_and_close_libreoffice(output_filename)
            print("Closed 'podium_highlighted.xlsx' window.")
            return selected_highlight, sel_val, False
        else:
            print(f"Error: Cell {selected_highlight} is not one of the highlighted Neighbours. Please rectify and try again.")


def step6_to_10_generate_pool_and_filter(sel_val: str) -> Set[str]:
    """Handles Steps 6 through 10: Generating Family, Blocks, Neighbours and Filtering."""
    print("\n--- Step 6: Family Generation ---")
    family_numbers = get_family(sel_val)
    for fn in family_numbers:
        print(fn)
        
    print("\n--- Step 7: Blocks for Family ---")
    all_generated_numbers = set()
    for fn in family_numbers:
        all_generated_numbers.add(fn)
        print(f"--- Block for Family Number {fn} ---")
        fn_block = get_block(fn)
        for key, val in fn_block.items():
            all_generated_numbers.add(val)
            print(f"{key:<10}: {val}")
        print()
        
    print("\n--- Step 8: Generating Block ---")
    b_val = sel_val
    b_block = get_block(b_val)
    for key, val in b_block.items():
        print(f"{key:<10}: {val}")
        
    all_generated_numbers.update(b_block.values())
    
    print("\n--- Step 9: Generating Neighbors for the 40-Member Pool ---")
    neighbour_pool = set()
    for num in sorted(list(all_generated_numbers)):
        nbs = get_neighbours(num)
        for nb in nbs:
            neighbour_pool.add(nb)
    print(f"Total unique Neighbors generated: {len(neighbour_pool)}")
            
    print("\n--- Step 10: 24-Permutation Removal Filter ---")
    final_highlights: Set[str] = set()
    removed_cnt: int = 0
    for nb in sorted(list(neighbour_pool)):
        if calculate_permutations(nb) == 24:
            removed_cnt += 1
        else:
            final_highlights.add(nb)
    
    print(f"Removed {removed_cnt} numbers with 24 permutations.")
    print(f"Retained {len(final_highlights)} numbers with repeating digits (4, 6, or 12 perms).")
    return final_highlights


def step11_highlight_final_neighbors(
    ws: Any, wb: Any, output_filename: str, final_highlights: Set[str], selected_highlight_coord: str, target_offset: int
) -> List[Tuple[str, str]]:
    """Handles Step 11: Highlighting final neighbors."""
    print(f"\n--- Step 11: Searching and Highlighting Final Neighbors ---")
    wait_and_close_libreoffice(
        output_filename, 
        custom_msg="Waiting for LibreOffice to release the lock on 'podium_highlighted.xlsx' before saving final red highlights..."
    )
    
    light_red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    found_red_cnt: int = 0
    red_matches_summary: List[Tuple[str, str]] = []
    
    ts_cell = cast(Any, ws)[selected_highlight_coord]
    target_os_row, target_os_col = get_block_origin(ws, ts_cell.row, ts_cell.column) # type: ignore
    print(f"Target Block Origin: Row {target_os_row}, Col {target_os_col}")
    
    for col_idx in range(4, 19):
        for row_idx in range(7, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip() != "":
                val_str = str(cell.value).strip().zfill(4)
                if val_str not in final_highlights:
                    continue

                m_start_row, m_start_col = get_block_origin(ws, row_idx, col_idx)
                m_offset = int(row_idx) - int(m_start_row)
                
                if (m_offset == target_offset) or ({m_offset, target_offset} == {0, 5}):
                    is_above = m_start_row < target_os_row
                    is_left_same_row = (m_start_row == target_os_row) and (m_start_col < target_os_col)
                    
                    if is_above or is_left_same_row:
                        cell.fill = light_red_fill
                        found_red_cnt += 1
                        red_matches_summary.append((cell.coordinate, val_str))
                        
    print(f"Found and highlighted {found_red_cnt} occurrence(s) in light red (Same Company, Left/Above only).")
    print("\nSaving final results to 'podium_highlighted.xlsx'...")
    save_workbook_safely(wb, output_filename)
            
    print("Done! Automatically opening the generated file...")
    open_file(output_filename)
            
    if red_matches_summary:
        print("\n--- Final Results (Light Red Highlights) ---")
        for coord, val in red_matches_summary:
            print(f"Cell {coord}: {val}")
            
    return red_matches_summary


def step12_select_red_cell(
    ws: Any, red_matches_summary: List[Tuple[str, str]], selected_highlight_coord: str
) -> Tuple[Optional[str], Optional[str]]:
    """Handles Step 12: User selection of a red highlighted cell."""
    print("\n--- Step 12: Select Red Highlighted Cell and Generate Neighbours ---")
    valid_red_cells = {coord: val for coord, val in red_matches_summary}

    if not valid_red_cells:
        print("No light-red highlighted cells were found; skipping Step 12-14.")
        return None, None

    ts_cell = cast(Any, ws)[selected_highlight_coord]
    target_os_row, target_os_col = get_block_origin(ws, ts_cell.row, ts_cell.column) # type: ignore

    step12_coord = input("\nEnter the cell coordinate of a light-red highlighted cell (or type 'done' to finish): ").strip().upper()
    user_provided = not (step12_coord == 'DONE' or step12_coord == 'EXIT' or not step12_coord)

    step12_val = None
    process_step12 = False

    if user_provided:
        if step12_coord not in valid_red_cells:
            print(f"Error: {step12_coord} is not in the red highlighted list. Will attempt auto-selection.")
        else:
            try:
                rc_cell = cast(Any, ws)[step12_coord]
                r_start_row, r_start_col = get_block_origin(ws, int(rc_cell.row), int(rc_cell.column))
                if r_start_row == target_os_row and r_start_col == target_os_col:
                    print("Selected red cell is from the same day as the yellow cell. Will attempt auto-selection.")
                else:
                    step12_val = valid_red_cells[step12_coord]
                    process_step12 = True
                    print(f"User-selected red highlighted cell {step12_coord} with value {step12_val}.")
            except Exception as e:
                print(f"Error validating selected red cell {step12_coord}: {e}. Will attempt auto-selection.")

    if not process_step12:
        print("No valid red-cell selection provided; auto-selection is disabled. Skipping Step 12 iterative processing.")
        return None, None

    if step12_val is not None:
        step12_neighbours = get_neighbours(step12_val)
        print(f"Neighbours for {step12_val}: {', '.join(step12_neighbours)}")
        close_file("podium_highlighted.xlsx")
        time.sleep(1)
        
    return step12_coord, step12_val


def step13_highlight_family_tree(
    sel_val: str, step12_val: str, script_dir: str
) -> Tuple[List[Tuple[str, str]], List[str]]:
    """Handles Step 13: Searching and Highlighting Neighbours in family_tree.xlsx."""
    print("\n--- Step 13: Searching and Highlighting Neighbours in family_tree.xlsx ---")
    family_tree_filename = os.path.join(script_dir, "family_tree.xlsx")
    family_tree_output_filename = os.path.join(script_dir, "family_tree_highlighted.xlsx")
    
    blue_matches_summary: List[Tuple[str, str]] = []
    
    if not os.path.exists(family_tree_filename):
        print(f"Error: Could not find '{family_tree_filename}' in the current directory.")
        print("Skipping Step 13.")
        return [], []
        
    print(f"Loading '{family_tree_filename}'... (this might take a few seconds)")
    try:
        ft_wb = openpyxl.load_workbook(family_tree_filename, data_only=False)
        ft_ws = cast(Any, ft_wb.active)
    except Exception as e:
        print(f"Error reading family_tree.xlsx: {e}")
        print("Skipping Step 13.")
        return [], []
        
    grey_fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
    light_blue_fill = PatternFill(start_color="FFCCFFFF", end_color="FFCCFFFF", fill_type="solid")
    found_grey_cnt: int = 0
    found_blue_cnt: int = 0
    grey_matches_summary: List[Tuple[str, str]] = []
    
    columns_with_step5: Set[int] = set()
    step5_val_str = str(sel_val).strip().zfill(4)
    for row_idx in range(3, 47):
        for col_idx in range(3, 26):
            cell = ft_ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip() != "":
                cell_val_str = str(cell.value).strip().zfill(4)
                if cell_val_str == step5_val_str:
                    cell.fill = grey_fill
                    found_grey_cnt += 1
                    grey_matches_summary.append((cell.coordinate, cell_val_str))
                    columns_with_step5.add(col_idx)
    
    print(f"Found and highlighted {found_grey_cnt} occurrence(s) of Step 5 value {step5_val_str} in grey.")
    print(f"Identified columns with Step 5 value: {sorted(columns_with_step5)}")
    
    step12_neighbours = get_neighbours(step12_val)
    print(f"Neighbours for {step12_val}: {', '.join(step12_neighbours)}")
    
    for row_idx in range(3, 47):
        for col_idx in range(3, 26):
            if col_idx in columns_with_step5:
                cell = ft_ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip() != "":
                    cell_val_str = str(cell.value).strip().zfill(4)
                    if cell_val_str in step12_neighbours:
                        cell.fill = light_blue_fill
                        found_blue_cnt += 1
                        blue_matches_summary.append((cell.coordinate, cell_val_str))
    
    print(f"Found and highlighted {found_blue_cnt} occurrence(s) of Neighbours in light blue.")
    print(f"Saving results to '{family_tree_output_filename}'...")
    save_workbook_safely(ft_wb, family_tree_output_filename)
            
    print("Done! Automatically opening the generated file...")
    open_file(family_tree_output_filename)
    
    if grey_matches_summary or blue_matches_summary:
        print("\n--- Grey and Blue Highlight Results ---")
        if grey_matches_summary:
            print("Grey highlights (Step 5 value):")
            for coord, val in grey_matches_summary:
                print(f"Cell {coord}: {val}")
        if blue_matches_summary:
            print("Blue highlights (Neighbours):")
            for coord, val in blue_matches_summary:
                print(f"Cell {coord}: {val}")

    unique_blue_values = sorted({val for _, val in blue_matches_summary})
    return blue_matches_summary, unique_blue_values


def step14_generate_and_highlight_next_podiums(
    output_filename: str, script_dir: str, unique_blue_values: List[str], green_cell_coord: str
) -> None:
    """Handles Step 14: Generate Blocks and Neighbours, then Highlight Duplicates."""
    print("\n--- Step 14: Generating Blocks and Neighbours, then Highlighting Duplicates ---")
    if not unique_blue_values:
        print("No unique blue values found; skipping Step 14 for this cell.")
        return

    color_map = ["#f2dceb", "#79d0f2", "#357826", "#b0c8e9"]
    next_podiums_filename = os.path.join(script_dir, "next_podiums.xlsx")
    
    try:
        ph_wb = openpyxl.load_workbook(output_filename, data_only=False)
        ph_ws = cast(Any, ph_wb.active)
    except Exception as e:
        print(f"Error re-loading podium_highlighted.xlsx: {e}")
        print("Skipping Step 14 due to file load error.")
        return

    # Remove all light red highlights
    for row in ph_ws.iter_rows(min_row=7, max_row=ph_ws.max_row, min_col=4, max_col=18):
        for cell in row:
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == "FFFF9999":
                cell.fill = PatternFill(fill_type=None)
    
    print("Cleared light red highlights from podium_highlighted.xlsx.")
    
    green_offset = get_company_offset(ph_ws, ph_ws[green_cell_coord].row, ph_ws[green_cell_coord].column)
    print(f"Green-highlighted cell: {green_cell_coord}, Company offset: {green_offset} ({COMPANIES.get(green_offset, 'Unknown')})")

    total_purple_matches = 0
    purple_matches_summary: List[Tuple[str, str, str]] = []

    for idx, blue_val in enumerate(unique_blue_values):
        active_color = color_map[idx % len(color_map)]
        fill = PatternFill(start_color=f"FF{active_color.lstrip('#')}", end_color=f"FF{active_color.lstrip('#')}", fill_type="solid")

        block_values = set(get_block(blue_val).values())
        neighbours_from_this_block = set()
        for num in block_values:
            neighbours_from_this_block.update(get_neighbours(num))

        print(f"Processing blue value #{idx+1} '{blue_val}': block members {sorted(block_values)}, neighbours {sorted(neighbours_from_this_block)}")

        matched_count = 0
        for row_idx in range(7, ph_ws.max_row + 1):
            for col_idx in range(4, 19):
                cell = ph_ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip() != "":
                    cell_val_str = str(cell.value).strip().zfill(4)
                    if cell_val_str in neighbours_from_this_block:
                        m_start_row, m_start_col = get_block_origin(ph_ws, row_idx, col_idx)
                        m_offset = int(row_idx) - int(m_start_row)
                        if (m_offset == green_offset) or ({m_offset, green_offset} == {0, 5}):
                            cell.fill = fill
                            matched_count += 1
                            total_purple_matches += 1
                            purple_matches_summary.append((cell.coordinate, cell_val_str, active_color))

        print(f"Highlighted {matched_count} cells for blue value '{blue_val}' using color {active_color}.")

    print(f"Found and highlighted {total_purple_matches} occurrence(s) of neighbours from blocks in color-coded highlights (company filter applied).")

    print(f"Saving results to '{next_podiums_filename}'...")
    save_workbook_safely(ph_wb, next_podiums_filename)

    print("Done! Attempting to open the generated files...")
    abs_next = os.path.abspath(next_podiums_filename)
    opened_next = open_file(abs_next)
    if not opened_next:
        print(f"Note: Could not open '{abs_next}' automatically. You can open it manually.")
    abs_out = os.path.abspath(output_filename)
    opened_out = open_file(abs_out)
    if not opened_out:
        print(f"Note: Could not open '{abs_out}' automatically. You can open it manually.")
    
    if purple_matches_summary:
        print("\n--- Four Colour Highlight Results ---")
        for coord, val, color in purple_matches_summary:
            print(f"Cell {coord}: {val} (color {color})")

    if unique_blue_values:
        print("\n--- Final Unique Blue Value -> Color Mapping ---")
        for idx, blue_val in enumerate(unique_blue_values[:4]):
            active_color = color_map[idx % len(color_map)]
            colored_val = get_terminal_color(blue_val, active_color)
            print(f"{idx+1}. {colored_val}")

    print("\n'next_podiums.xlsx' and updated 'podium_highlighted.xlsx' are now open for your review.")


def main():
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_filename = os.path.join(_script_dir, "podium.xlsx")
    output_filename = os.path.join(_script_dir, "podium_highlighted.xlsx")
    
    if not os.path.exists(excel_filename):
        print(f"Error: Could not find '{excel_filename}' in the current directory.")
        sys.exit(1)
        
    print(f"Loading '{excel_filename}'... (this might take a few seconds)")
    try:
        wb = openpyxl.load_workbook(excel_filename, data_only=False)
        ws = cast(Any, wb.active)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        sys.exit(1)

    while True:
        cell_coord, number, target_offset, company_name = step1_and_2_get_initial_cell(ws, excel_filename)
        
        neighbours = step3_and_4_highlight_neighbours(
            ws, wb, output_filename, cell_coord, number, target_offset, company_name
        )
        
        selected_highlight_coord, sel_val, user_exited = step5_select_highlighted_cell(
            ws, output_filename, neighbours, target_offset, cell_coord
        )
        if user_exited:
            break
            
        final_highlights = step6_to_10_generate_pool_and_filter(sel_val)
        
        red_matches_summary = step11_highlight_final_neighbors(
            ws, wb, output_filename, final_highlights, selected_highlight_coord, target_offset
        )
        
        step12_coord, step12_val = step12_select_red_cell(ws, red_matches_summary, selected_highlight_coord)
        
        if step12_val is not None and step12_coord is not None:
            blue_matches_summary, unique_blue_values = step13_highlight_family_tree(
                sel_val, step12_val, _script_dir
            )
            
            print(f"\n[Finished processing results for cell {step12_coord}]")
            
            step14_generate_and_highlight_next_podiums(
                output_filename, _script_dir, unique_blue_values, cell_coord
            )
            
        break

if __name__ == "__main__":
    try:
        main()
    finally:
        # Safely shut down the custom logger.
        stdout_obj = cast(Any, sys.stdout)
        if hasattr(stdout_obj, "close_log"):
            stdout_obj.close_log()

